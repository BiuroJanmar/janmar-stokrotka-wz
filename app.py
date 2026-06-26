import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from io import BytesIO
import re

st.set_page_config(page_title="Janmar WZ Stokrotka", layout="centered")

st.title("JANMAR WZ-Stokrotka Web v8.6")
st.subheader("Oficjalny, ultra-stabilny generator dokumentów WZ (Silnik Eurocash)")
st.write("Wgraj surowy plik tekstowy (.TXT) zamówienia ze Stokrotki.")

PRZELICZNIKI_SIECI = {
    "KALAFIOR": 6.0, "KAPUSTA PEKIŃSKA": 10.0, "KAPUSTA BIAŁA": 10.0,
    "KAPUSTA CZERWONA": 10.0, "KAPUSTA WŁOSKA": 10.0, "KAPUSTA WCZESNA": 6.0,
    "CUKINIA": 5.0, "KALAREPA": 8.0, "KUKURYDZA": 30.0, "RZODKIEWKA": 10.0,
    "SAŁATA LODOWA": 10.0, "SAŁATA MASŁOWA": 12.0, "SELER NACIOWY": 16.0, "MARCHEW": 10.0
}

def ustal_kraj_pochodzenia(nazwa_towaru):
    nazwa_upper = nazwa_towaru.upper()
    if any(x in nazwa_upper for x in ["POMARAŃCZE", "MANDARYNKI"]): return "HISZPANIA / TURCJA"
    if "ZIEMNIAKI IMPORTOWE" in nazwa_upper: return "CYPR / GRECJA"
    if "ARBUZ" in nazwa_upper: return "WŁOCHY / HISZPANIA"
    return "POLSKA"

uploaded_file = st.file_uploader("Wgraj plik zamówienia Stokrotki (.TXT)", type=["txt"])

if uploaded_file is not None:
    nazwa_pliku = uploaded_file.name
    file_bytes = uploaded_file.read()
    
    tekst = None
    for enc in ['cp1250', 'windows-1250', 'utf-8', 'latin1']:
        try:
            tekst = file_bytes.decode(enc)
            if "ZAMÓWIENIE" in tekst or "Stokrotka" in tekst or "Termin" in tekst:
                break
        except:
            continue
            
    if tekst is None:
        tekst = file_bytes.decode('utf-8', errors='ignore')
    
    linie = tekst.splitlines()
    
    nr_zam = "Nieznany"
    for line in linie:
        if "ZAMÓWIENIE TOWARU NR" in line.upper():
            try:
                nr_zam = line.upper().split("NR")[1].strip().split()[0]
            except:
                pass
                
    if nr_zam == "Nieznany" or not nr_zam:
        nazwa_pliku_lower = nazwa_pliku.lower()
        if "nr" in nazwa_pliku_lower:
            try: nr_zam = nazwa_pliku_lower.split("nr")[1].split()[0].replace(".txt","")
            except: nr_zam = nazwa_pliku.replace(".txt","").replace(".TXT","")
        else:
            nr_zam = nazwa_pliku.replace(".txt","").replace(".TXT","")
            
    nr_zam = "".join([c for c in nr_zam if c.isdigit() or c in ['/', '_', '-']])

    towary = []
    for line_raw in linie:
        line_clean = line_raw.strip()
        if not line_clean:
            continue
            
        # SZUKAMY KODU TOWARU NA POCZĄTKU LINII (DOKŁADNIE JAK W EUROCASH)
        match_kod = re.match(r'^(\d{6})', line_clean)
        if match_kod:
            kod_part = match_kod.group(1)
            
            try:
                # Szukamy wszystkich ciągów liczb/ilości w linii
                liczby = re.findall(r'\b\d+[\.,]\d+\b|\b\d+\b', line_clean)
                if len(liczby) >= 2:
                    # Ostatnia liczba to ilość zamówiona (odrzucamy ewentualny EAN przed nią)
                    ilosc_str = liczby[-1]
                    ilosc_koncowa = float(ilosc_str.replace(',', '.'))
                    
                    # Wyciągamy jednostkę miary (szt/kg)
                    jm = "szt"
                    if "kg" in line_clean.lower():
                        jm = "kg"
                        
                    # Wycinamy nazwę produktu pomiędzy kodem a resztą danych
                    środek = line_clean[6:].strip()
                    # Usuwamy z nazwy kody EAN i końcowe liczby, żeby została czysta nazwa towaru
                    nazwa_towaru = re.sub(r'\s+\d{10,14}\s+.*|\s+\d+[\.,]\d+\s*.*|\s+\d+\s*$', '', środek).strip().upper()
                    
                    # Jeśli nazwa ucięła się za mocno, bierzemy prosty fallback bezpieczny dla oka
                    if len(nazwa_towaru) < 2:
                        nazwa_towaru = "TOWAR " + kod_part
                        
                    if ilosc_koncowa > 0:
                        w_opak = 10.0
                        for klucz, waga in PRZELICZNIKI_SIECI.items():
                            if klucz in nazwa_towaru:
                                w_opak = waga
                                break
                                
                        ilosc_op = round(ilosc_koncowa / w_opak, 1) if w_opak > 0 else 0
                        kraj = ustal_kraj_pochodzenia(nazwa_towaru)
                        
                        towary.append({
                            "kod": kod_part, "nazwa": nazwa_towaru, "jm": jm,
                            "w_opak": w_opak, "ilosc_op": ilosc_op, "ilosc_koncowa": ilosc_koncowa,
                            "kraj": kraj
                        })
            except:
                pass

    if towary:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dokument WZ"
        ws.views.sheetView[0].showGridLines = True
        
        font_title = Font(name="Arial", size=15, bold=True, color="1F497D")
        font_section = Font(name="Arial", size=10, bold=True, color="000000")
        font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        font_body = Font(name="Arial", size=10, bold=False, color="000000")
        font_body_bold = Font(name="Arial", size=10, bold=True, color="000000")
        
        header_fill = PatternFill(start_color="1F497D", fill_type="solid")
        info_bar_fill = PatternFill(start_color="E9EDF4", fill_type="solid")
        zebra_fill = PatternFill(start_color="F2F5F8", fill_type="solid")
        
        thin = Side(style='thin', color='D9D9D9')
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
        double_bottom = Border(top=Side(style='thin', color='D9D9D9'), bottom=Side(style='double', color='1F497D'))
        
        ws['A1'] = "DOKUMENT WZ"
        ws['A1'].font = font_title
        ws.merge_cells('G1:H1')
        ws['G1'] = f"Nr WZ: STOK/{nr_zam}"
        ws['G1'].font = font_body_bold
        ws['G1'].alignment = Alignment(horizontal="right")
        
        ws['D3'] = "DOSTAWCA:"
        ws['D3'].font = font_section
        ws['D4'] = "GPW JANMAR SP. Z O.O. SP. K."
        ws['D4'].font = font_body_bold
        ws['D5'] = "ul. Gołaśka 3/58, 30-619 Kraków"
        ws['D5'].font = font_body
        
        ws['G3'] = "ODBIORCA:"
        ws['G3'].font = font_section
        ws['G4'] = "STOKROTKA SP. Z O.O."
        ws['G4'].font = font_body_bold
        ws['G5'] = "ul. Projektowa 1, 20-209 Lublin"
        ws['G5'].font = font_body
        
        data_zamowienia = "................"
        data_dostawy = "................"
        miejsce_dostawy = "STOKROTKA CD"
        
        for line in linie:
            if "Termin dostawy" in line:
                match_dost = re.search(r'Termin dostawy\s+([\d\.]+)', line)
                match_wyst = re.search(r'Data wystawienia\s+([\d\.]+)', line)
                if match_dost: data_dostawy = match_dost.group(1)
                if match_wyst: data_zamowienia = match_wyst.group(1)
            if "Towar należy dostarczyć:" in line:
                try:
                    idx = linie.index(line)
                    if idx + 1 < len(linie): miejsce_dostawy = linie[idx+1].strip()
                except: pass

        ws.cell(row=7, column=1, value="Nr zamówienia Stokrotka:").font = font_body_bold
        ws.cell(row=7, column=3, value=nr_zam).font = font_body
        ws.cell(row=8, column=1, value="Data zamówienia:").font = font_body_bold
        ws.cell(row=8, column=3, value=data_zamowienia).font = font_body
        ws.cell(row=9, column=1, value="Data dostawy:").font = font_body_bold
        ws.cell(row=9, column=3, value=data_dostawy).font = font_body
        
        ws.merge_cells('A11:H11')
        ws['A11'] = f" MIEJSCE DOSTAWY: {miejsce_dostawy.upper()}"
        ws['A11'].font = font_body_bold
        ws['A11'].fill = info_bar_fill
        ws['A11'].alignment = Alignment(vertical="center")
        ws.row_dimensions[11].height = 24
        
        naglowki = ["Lp.", "Kod towaru", "Nazwa asortymentu", "Kraj pochodzenia", "Jm.", "W opak.", "Ilość op.", "Ilość szt./kg"]
        for col_idx, text in enumerate(naglowki, start=1):
            cell = ws.cell(row=14, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[14].height = 24
            
        start_row = 15
        max_wiersz_siatki = 60
        
        for r in range(start_row, max_wiersz_siatki + 1):
            idx_towaru = r - start_row
            ws.row_dimensions[r].height = 22
            ws.cell(row=r, column=1, value=idx_towaru + 1).alignment = Alignment(horizontal="center")
            
            for c in range(1, 9):
                cell = ws.cell(row=r, column=c)
                cell.border = border_all
                cell.font = font_body
                if idx_towaru % 2 == 1: cell.fill = zebra_fill
                    
            if idx_towaru < len(towary):
                t = towary[idx_towaru]
                ws.cell(row=r, column=2, value=t['kod']).alignment = Alignment(horizontal="center")
                ws.cell(row=r, column=3, value=t['nazwa']).alignment = Alignment(horizontal="left")
                ws.cell(row=r, column=4, value=t['kraj']).alignment = Alignment(horizontal="center")
                ws.cell(row=r, column=5, value=t['jm']).alignment = Alignment(horizontal="center")
                
                c_w_opak = ws.cell(row=r, column=6, value=t['w_opak'])
                c_w_opak.alignment = Alignment(horizontal="right")
                c_w_opak.number_format = '#,##0.0'
                
                c_op = ws.cell(row=r, column=7, value=t['ilosc_op'])
                c_op.alignment = Alignment(horizontal="right")
                c_op.number_format = '#,##0.0'
                
                c_koncowa = ws.cell(row=r, column=8, value=t['ilosc_koncowa'])
                c_koncowa.alignment = Alignment(horizontal="right")
                c_koncowa.number_format = '#,##0'
            else:
                ws.cell(row=r, column=4, value="")

        sum_row = max_wiersz_siatki + 2
        ws.row_dimensions[sum_row].height = 24
        ws.cell(row=sum_row, column=6, value="RAZEM NETTO:").font = font_body_bold
        ws.cell(row=sum_row, column=6).alignment = Alignment(horizontal="right")
        
        sum_cell = ws.cell(row=sum_row, column=8, value=f"=SUM(H15:H{max_wiersz_siatki})")
        sum_cell.font = font_body_bold
        sum_cell.alignment = Alignment(horizontal="right")
        sum_cell.border = double_bottom
        sum_cell.number_format = '#,##0'
        
        footer_row = sum_row + 2
        ws.row_dimensions[footer_row].height = 22
        ws.cell(row=footer_row, column=1, value="DOKUMENT SPORZĄDZIŁ: .............................").font = font_body_bold
        ws.cell(row=footer_row, column=6, value="ILOŚĆ PALET EURO: ..................").font = font_body_bold
        ws.cell(row=footer_row, column=6).alignment = Alignment(horizontal="right")

        kraje_lista = '"POLSKA, HISZPANIA, HOLANDIA, PORTUGALIA, WŁOCHY, GRECJA, NIEMCY, FRANCJA, TURCJA, MAROKO"'
        dv = DataValidation(type="list", formula1=kraje_lista, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"D15:D{max_wiersz_siatki}")
        
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 13
        ws.column_dimensions['C'].width = 38
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 14
        
        output = BytesIO()
        wb.save(output)
        wz_excel = output.getvalue()
        
        st.success(f"Sukces! Wygenerowano pełną WZ Stokrotka dla zamówienia nr: {nr_zam}")
        st.download_button(
            label="📥 Pobierz oficjalną WZ Stokrotka (Excel)",
            data=wz_excel,
            file_name=f"WZ_STOKROTKA_{nr_zam}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.error("Błąd odczytu. System nie odnalazł linii produktowych.")
